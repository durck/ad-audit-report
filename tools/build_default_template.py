"""Build the bundled default template from scratch — zero traces of any client data.

Generates src/adreport/templates/default.xlsx with:
  - Sheet "Результаты": 9-column finding table, drop-down validations on D/E
  - Sheet "Справочники": list-source values for the drop-downs
  - No appendix sheets (created dynamically by render_report)
  - No images, no printer settings, no example data
  - No author metadata in core.xml

Run once when the template structure needs to change. Output is committed to git.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = Path(__file__).resolve().parents[1] / "src" / "adreport" / "templates" / "default.xlsx"

HEADERS = [
    "№",
    "Дата",
    "Объект тестирования",
    "Сегмент сети, из которого осуществлялась проверка",
    "Тип",
    "Описание",
    "Подробности",
    "Предлагаемые мероприятия",
    "Примечания",
]

SEGMENT_VALUES = ["Пользовательский", "Серверный", "Телефонный", "Принтерный", "-"]
TYPE_VALUES = ["Недостаток", "Возможно недостаток", "Уязвимость"]

LAST_DATA_ROW = 200  # data validation / table range


def build() -> None:
    wb = openpyxl.Workbook()
    main = wb.active
    main.title = "Результаты"
    refs = wb.create_sheet("Справочники")

    _build_main_sheet(main)
    _build_refs_sheet(refs)
    _add_validations(main)

    # Clear default openpyxl metadata that would otherwise show as the author.
    wb.properties.creator = ""
    wb.properties.lastModifiedBy = ""
    wb.properties.title = ""

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote clean template: {OUT}")
    print(f"Size: {OUT.stat().st_size} bytes")


def _build_main_sheet(ws) -> None:
    # Title in merged A3:I4
    ws["A3"] = "Список выявленных недостатков информационной безопасности"
    ws["A3"].font = Font(bold=True, size=14)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells("A3:I4")

    # Column widths
    widths = [5, 12, 22, 22, 18, 35, 30, 40, 25]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[5].height = 45

    # Default per-column data format (applied to any cell in the column that
    # doesn't override `s`):
    #   - column B (Дата) → date format DD.MM.YYYY
    #   - all other columns inherit wrap-text + top-vertical via the table style
    ws.column_dimensions["B"].number_format = "DD.MM.YYYY"

    # Header row at r5
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, name in enumerate(HEADERS, start=1):
        c = ws.cell(row=5, column=col, value=name)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = border

    # Pre-populate rows 6 and 7 as styled "sample" cells. The renderer reads
    # per-column `s` (style id) from these samples via _template_style_for_column.
    # Row 6 = regular data style.
    # Row 7 = hyperlink data style (only column G — used when a cell points
    #         at an appendix sheet).
    #
    # We use inlineStr placeholders ("") rather than openpyxl's default
    # `<c t="n"/>` so the cells are well-formed (malformed empty number cells
    # would make Excel reject the workbook).
    data_align = Alignment(wrap_text=True, vertical="top")
    for col in range(1, len(HEADERS) + 1):
        c = ws.cell(row=6, column=col, value="")
        c.alignment = data_align
        c.border = border
    ws.cell(row=6, column=2).number_format = "DD.MM.YYYY"

    # Row 7 G: hyperlink style — blue underlined font, otherwise same as data
    hyperlink_font = Font(color="0563C1", underline="single")
    g7 = ws.cell(row=7, column=7, value="")
    g7.alignment = data_align
    g7.border = border
    g7.font = hyperlink_font

    # Row dimensions: row 6/7 stay default height, otherwise Excel auto-fits.
    # Row 7 is a private sample only — hide it so the user never sees the
    # empty cell.
    ws.row_dimensions[7].hidden = True

    # Excel Table on the data range
    ref = f"A5:{get_column_letter(len(HEADERS))}{LAST_DATA_ROW}"
    tbl = Table(displayName="Findings", ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight13",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)


def _build_refs_sheet(ws) -> None:
    ws["A1"] = "Сегмент сети"
    ws["C1"] = "Тип"
    bold = Font(bold=True)
    ws["A1"].font = bold
    ws["C1"].font = bold

    for i, v in enumerate(SEGMENT_VALUES, start=2):
        ws.cell(row=i, column=1, value=v)
    for i, v in enumerate(TYPE_VALUES, start=2):
        ws.cell(row=i, column=3, value=v)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["C"].width = 24


def _add_validations(ws) -> None:
    # Drop-down for column D (Сегмент сети) — uses list from "Справочники"!A2:A6
    seg_last = 1 + len(SEGMENT_VALUES)
    seg_dv = DataValidation(
        type="list",
        formula1=f"=Справочники!$A$2:$A${seg_last}",
        allow_blank=True,
        showErrorMessage=True,
    )
    seg_dv.add(f"D6:D{LAST_DATA_ROW}")
    ws.add_data_validation(seg_dv)

    # Drop-down for column E (Тип) — uses list from "Справочники"!C2:C4
    type_last = 1 + len(TYPE_VALUES)
    type_dv = DataValidation(
        type="list",
        formula1=f"=Справочники!$C$2:$C${type_last}",
        allow_blank=True,
        showErrorMessage=True,
    )
    type_dv.add(f"E6:E{LAST_DATA_ROW}")
    ws.add_data_validation(type_dv)


if __name__ == "__main__":
    build()
