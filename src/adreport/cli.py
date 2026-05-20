"""Typer CLI: adreport build|init-config|validate|list-rules."""

from __future__ import annotations

import sys
from pathlib import Path

import typer

from .catalog import Catalog
from .config import DEFAULT_PROJECT_YAML, ProjectConfig
from .parsers import PlumHoundLoader, load_pingcastle_details, parse_pingcastle
from .pipeline import build_findings
from .renderer import render_report
from .sanitize import sanitize_template

app = typer.Typer(add_completion=False, help="Generate Excel pentest reports from PingCastle + PlumHound.")


def _resolve_relative(base: Path, p: Path) -> Path:
    """Resolve paths in project.yaml relative to the config file's directory."""
    return p if p.is_absolute() else (base.parent / p).resolve()


def _resolve_domain_input(base: Path, d):
    """Resolve relative paths inside a DomainInput against the config dir."""
    pc = _resolve_relative(base, d.pingcastle) if d.pingcastle else None
    html = _resolve_relative(base, d.pingcastle_html) if d.pingcastle_html else None
    if html is None and pc is not None:
        cand = pc.with_suffix(".html")
        if cand.exists():
            html = cand
    plum = _resolve_relative(base, d.plumhound) if d.plumhound else None
    return d.model_copy(update={"pingcastle": pc, "pingcastle_html": html, "plumhound": plum})


def _load_project(path: Path) -> tuple[ProjectConfig, Path]:
    """Load ProjectConfig and return alongside the config's parent directory for relative-path resolution."""
    cfg = ProjectConfig.load(path)
    base = path.resolve()

    update: dict = {
        "template": _resolve_relative(base, cfg.template) if cfg.template is not None else None,
        "output": _resolve_relative(base, cfg.output),
    }

    if cfg.inputs is not None:
        pingcastle_path = _resolve_relative(base, cfg.inputs.pingcastle)
        html_path = (
            _resolve_relative(base, cfg.inputs.pingcastle_html) if cfg.inputs.pingcastle_html else None
        )
        if html_path is None:
            candidate = pingcastle_path.with_suffix(".html")
            if candidate.exists():
                html_path = candidate
        update["inputs"] = cfg.inputs.model_copy(
            update={
                "pingcastle": pingcastle_path,
                "pingcastle_html": html_path,
                "plumhound": _resolve_relative(base, cfg.inputs.plumhound) if cfg.inputs.plumhound else None,
            }
        )
    if cfg.domains:
        update["domains"] = [_resolve_domain_input(base, d) for d in cfg.domains]

    cfg = cfg.model_copy(update=update)
    return cfg, base.parent


@app.command()
def init_config(path: Path = typer.Argument(Path("project.yaml"), help="Where to write the config file.")):
    """Write a starter project.yaml to PATH."""
    if path.exists():
        typer.secho(f"{path} already exists — refusing to overwrite.", fg=typer.colors.RED, err=True)
        raise typer.Exit(1)
    path.write_text(DEFAULT_PROJECT_YAML, encoding="utf-8")
    typer.secho(f"Wrote {path}", fg=typer.colors.GREEN)


def _run_pipeline_for_domains(cfg, catalog) -> tuple[list, list[str]]:
    """Run pipeline across every domain in cfg; return (findings, unknown_risk_ids).

    Prints a per-domain summary table with a legend so the meaning of each
    column is self-evident.

    Sources are independent: a domain may have only PingCastle, only PlumHound,
    or both. Missing PingCastle → PlumHound-only synthetic findings still fire.
    Missing both → domain is skipped with a warning.
    """
    from .model import PingCastleReport

    all_findings: list = []
    all_unknown: list[str] = []
    rows: list[dict] = []  # per-domain summary for the table

    for d in cfg.iter_domain_inputs():
        pc_path = d.pingcastle if d.pingcastle and d.pingcastle.exists() else None
        plum_path = d.plumhound if d.plumhound and d.plumhound.exists() else None
        if pc_path is None and plum_path is None:
            typer.secho(
                f"⚠ Skipping domain {d.name or '<single>'}: neither PingCastle XML nor PlumHound output found",
                fg=typer.colors.YELLOW,
            )
            continue

        if d.pingcastle and pc_path is None:
            typer.secho(
                f"  ℹ [{d.name}] PingCastle XML configured but not found at "
                f"{d.pingcastle} — falling back to PlumHound-only synthetics",
                fg=typer.colors.YELLOW,
            )
        if d.plumhound and plum_path is None:
            typer.secho(
                f"  ℹ [{d.name}] PlumHound output configured but not found at "
                f"{d.plumhound} — appendix details will be limited to PingCastle HTML",
                fg=typer.colors.YELLOW,
            )

        if pc_path is not None:
            pc = parse_pingcastle(pc_path)
        else:
            pc = PingCastleReport(
                domain=d.name, generation_date="", global_score=0,
                stale_objects_score=0, privileged_group_score=0,
                trust_score=0, anomaly_score=0, rules=(),
            )

        pc_details = load_pingcastle_details(d.pingcastle_html) if d.pingcastle_html and d.pingcastle_html.exists() else {}
        plum = PlumHoundLoader(plum_path) if plum_path else None
        try:
            result = build_findings(
                pc, plum, catalog, cfg, pingcastle_details=pc_details, domain=d.name
            )
        finally:
            if plum is not None:
                plum.cleanup()

        all_findings.extend(result.findings)
        all_unknown.extend(result.unknown_risk_ids)
        rows.append({
            "domain": d.name or pc.domain or "<single>",
            "pc_rules": len(pc.rules) if pc_path else None,
            "pc_score": pc.global_score if pc_path else None,
            "html_det": len(pc_details) if pc_path else None,
            "plum": plum is not None,
            "findings": len(result.findings),
        })

    if rows:
        _print_domain_summary_table(rows)
    return all_findings, sorted(set(all_unknown))


def _print_domain_summary_table(rows: list[dict]) -> None:
    """Render the per-domain breakdown as a legend + aligned table."""
    typer.echo("")
    typer.secho("Легенда:", fg=typer.colors.CYAN, bold=True)
    typer.echo("  PC_rules    — правил PingCastle, найденных в .xml")
    typer.echo("  PC_score    — GlobalScore PingCastle (0..100, выше = хуже)")
    typer.echo("  HTML_det    — правил с детализацией в PingCastle HTML")
    typer.echo("  Plum        — PlumHound подключён (✓/—)")
    typer.echo("  Findings    — строк в финальном отчёте")
    typer.echo("")

    headers = ("Домен", "PC_rules", "PC_score", "HTML_det", "Plum", "Findings")
    # Compute width per column from header + max value
    def _fmt(v):
        if v is None:
            return "—"
        if isinstance(v, bool):
            return "✓" if v else "—"
        return str(v)

    table_rows = [
        (r["domain"], _fmt(r["pc_rules"]), _fmt(r["pc_score"]),
         _fmt(r["html_det"]), _fmt(r["plum"]), _fmt(r["findings"]))
        for r in rows
    ]
    widths = [
        max(len(headers[i]), *(len(row[i]) for row in table_rows))
        for i in range(len(headers))
    ]
    # Left-align domain, right-align numbers
    align = ("<", ">", ">", ">", "^", ">")

    def _row(cells):
        return "  " + "  ".join(
            f"{cells[i]:{align[i]}{widths[i]}}" for i in range(len(cells))
        )

    sep = "  " + "─" * (sum(widths) + 2 * (len(widths) - 1))
    typer.secho(_row(headers), fg=typer.colors.CYAN, bold=True)
    typer.echo(sep)
    for row in table_rows:
        typer.echo(_row(row))
    typer.echo(sep)
    typer.secho(
        f"Итого: {len(rows)} доменов, {sum(r['findings'] for r in rows)} строк в отчёте",
        fg=typer.colors.GREEN, bold=True,
    )


@app.command()
def validate(
    project: Path = typer.Argument(..., help="Path to project.yaml."),
):
    """Parse the inputs, show coverage stats — but do not render the xlsx."""
    cfg, _ = _load_project(project)
    catalog = Catalog.load_default()
    findings, unknown = _run_pipeline_for_domains(cfg, catalog)
    typer.secho(f"Total findings prepared: {len(findings)}", fg=typer.colors.GREEN)
    if unknown:
        typer.secho(
            f"⚠ Unknown RiskIds: {', '.join(unknown)}",
            fg=typer.colors.YELLOW,
        )
        typer.echo("  → add entries for them to recommendations.yaml.")


@app.command()
def list_rules(
    project: Path | None = typer.Argument(None, help="Optional project.yaml — if given, marks rules present in this scan."),
):
    """List all RiskIds known to the catalog, optionally marking those in a specific scan."""
    catalog = Catalog.load_default()
    present: set[str] = set()
    extra: list[str] = []
    if project is not None:
        cfg, _ = _load_project(project)
        for d in cfg.iter_domain_inputs():
            if d.pingcastle.exists():
                pc = parse_pingcastle(d.pingcastle)
                present.update(r.risk_id for r in pc.rules)
        extra = sorted(present - set(catalog.recommendations))

    typer.secho(
        f"Catalog: {len(catalog.recommendations)} PingCastle rule(s), "
        f"{len(catalog.synthetic_findings)} synthetic finding(s)",
        fg=typer.colors.CYAN,
    )
    for risk_id in sorted(catalog.recommendations):
        marker = " [match]" if risk_id in present else ""
        typer.echo(f"  {risk_id}{marker}")
    if extra:
        typer.secho(f"\nIn this scan but missing from catalog:", fg=typer.colors.YELLOW)
        for risk_id in extra:
            typer.echo(f"  {risk_id}")


@app.command("sanitize-template")
def sanitize_template_cmd(
    template: Path = typer.Argument(..., help="Path to the corporate template xlsx (input)."),
    output: Path = typer.Argument(..., help="Where to write the sanitised template."),
    title: str = typer.Option(
        "Список недостатков конфигурации",
        "--title",
        help="Replacement for cell A3 (the report title that often contains the auditor's name).",
    ),
):
    """Strip confidential metadata from an xlsx template.

    Removes: creator / lastModifiedBy from docProps, absPath of the original
    author's filesystem from workbook.xml, Company / Manager from app.xml, and
    replaces the cell A3 report title with a neutral string.

    Preserves: all sheet structure, styles, data validations, named ranges,
    tables, drawings, printerSettings — i.e. the xlsx remains a usable template.
    """
    changes = sanitize_template(template, output, new_title=title)
    typer.secho(f"✓ Wrote {output}", fg=typer.colors.GREEN)
    if changes:
        typer.echo("Sanitised fields:")
        for k, v in changes.items():
            typer.echo(f"  {k}:  {v}")
    else:
        typer.secho("No confidential metadata found — template was already clean.", fg=typer.colors.YELLOW)


@app.command()
def repair(
    xlsx: Path = typer.Argument(..., help="Path to the suspect xlsx."),
    out: Path = typer.Argument(..., help="Where to write the openpyxl-roundtripped copy."),
):
    """Round-trip a built report through openpyxl's serialiser.

    openpyxl emits Microsoft-shaped XML (different attribute ordering, namespace
    placement, element flatness vs lxml's pretty output). If Excel rejects the
    original but accepts the round-tripped copy, the root cause is lxml-vs-Excel
    serialiser quirks, not our data.

    This is a diagnostic — the round-trip may strip our custom appendices,
    hyperlinks, or x14 data validations. Don't ship the output as the report;
    use it only to identify whether the lxml output is the problem.
    """
    try:
        import openpyxl as _opx
    except ImportError:
        typer.secho("openpyxl is not installed — `pip install openpyxl`", fg=typer.colors.RED, err=True)
        raise typer.Exit(1)

    if not xlsx.exists():
        typer.secho(f"File not found: {xlsx}", fg=typer.colors.RED, err=True)
        raise typer.Exit(1)

    typer.echo(f"Loading {xlsx}...")
    wb = _opx.load_workbook(xlsx)
    typer.echo(f"Loaded {len(wb.sheetnames)} sheets: {wb.sheetnames[:5]}{'...' if len(wb.sheetnames) > 5 else ''}")
    typer.echo(f"Saving roundtripped copy to {out}...")
    wb.save(out)
    typer.secho(
        f"✓ Wrote {out} ({out.stat().st_size} bytes)\n"
        f"  Try opening this file in Excel. If it opens cleanly while the original"
        f"  does not, the lxml serialiser is at fault (custom appendices may be"
        f"  lost in the round-trip — this is a diagnostic, not a final report).",
        fg=typer.colors.GREEN,
    )


@app.command()
def doctor(
    xlsx: Path = typer.Argument(..., help="Path to a built report xlsx to validate."),
    through_openpyxl: bool = typer.Option(
        False, "--through-openpyxl",
        help="Also load + re-save through openpyxl to surface strict-mode warnings.",
    ),
):
    """Sanity-check a generated xlsx for problems Excel will reject on open.

    Layer 1 (always):
      - well-formedness of every internal XML part
      - cell text length ≤ 32767 chars (Excel hard limit)
      - no XML 1.0 forbidden control characters in cell content
      - workbook ↔ rels ↔ ContentTypes consistency (sheetId / rId / Override)
      - no duplicate sheet names / sheetIds / rIds
      - per-sheet: rows in ascending `r` order, cells in ascending column order
      - per-sheet: no duplicate `r` attributes on rows or cells
      - per-sheet: hyperlinks point at existing locations (Sheet!Cell format)
      - per-sheet: <dimension ref="..."> sane

    Layer 2 (opt-in `--through-openpyxl`):
      - load the file via openpyxl in strict mode and surface any warnings
        (this catches things our text-level scanners miss — drawing rels,
        named-range references, table ranges past sheet bounds, etc.)

    Use this when Excel shows «Ошибка в части содержимого» on open and the
    standard render path produces a file that openpyxl loads but Excel rejects.
    """
    import re as _re
    import zipfile as _zip

    from lxml import etree as _etree

    NS_M = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    NS_PKG = "{http://schemas.openxmlformats.org/package/2006/relationships}"
    NS_REL_ = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    NS_CT_ = "{http://schemas.openxmlformats.org/package/2006/content-types}"
    INVALID_RE = _re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

    if not xlsx.exists():
        typer.secho(f"File not found: {xlsx}", fg=typer.colors.RED, err=True)
        raise typer.Exit(1)

    issues: list[str] = []
    over_limit: list[tuple[str, int]] = []
    bad_chars: list[str] = []

    with _zip.ZipFile(xlsx) as z:
        names = z.namelist()

        for n in names:
            if not (n.endswith(".xml") or n.endswith(".rels")):
                continue
            try:
                _etree.fromstring(z.read(n))
            except _etree.XMLSyntaxError as e:
                issues.append(f"XML parse error in {n}: {e}")

        # Load core parts up front — per-sheet checks need to know the list
        # of valid sheet names (for hyperlinks → location target checks).
        try:
            wb = _etree.fromstring(z.read("xl/workbook.xml"))
            rels = _etree.fromstring(z.read("xl/_rels/workbook.xml.rels"))
            ct = _etree.fromstring(z.read("[Content_Types].xml"))
        except KeyError as e:
            issues.append(f"Missing core part: {e}")
            wb = rels = ct = None

        cell_ref_re = _re.compile(r"^([A-Z]+)(\d+)$")

        def _col_index(letters: str) -> int:
            idx = 0
            for ch in letters:
                idx = idx * 26 + (ord(ch) - ord("A") + 1)
            return idx

        for n in names:
            if "worksheets/sheet" not in n or not n.endswith(".xml"):
                continue
            root = _etree.fromstring(z.read(n))

            # Cell content checks
            for t in root.iter(NS_M + "t"):
                if t.text is None:
                    continue
                if len(t.text) > 32767:
                    over_limit.append((n, len(t.text)))
                if INVALID_RE.search(t.text):
                    bad_chars.append(n)

            # Per-sheet structural checks
            sd = root.find(NS_M + "sheetData")
            if sd is None:
                continue

            row_numbers_seen: list[int] = []
            for row in sd.findall(NS_M + "row"):
                r_attr = row.get("r")
                if r_attr is None:
                    issues.append(f"{n}: <row> without r attribute")
                    continue
                rn = int(r_attr)
                if row_numbers_seen and rn <= row_numbers_seen[-1]:
                    issues.append(
                        f"{n}: row r='{rn}' not in ascending order "
                        f"(previous was r='{row_numbers_seen[-1]}')"
                    )
                row_numbers_seen.append(rn)

                cell_cols_seen: list[int] = []
                cell_refs_seen: set[str] = set()
                for c in row.findall(NS_M + "c"):
                    cref = c.get("r", "")
                    m = cell_ref_re.match(cref)
                    if not m:
                        issues.append(f"{n}: cell with invalid r='{cref}' in row {rn}")
                        continue
                    if cref in cell_refs_seen:
                        issues.append(f"{n}: duplicate cell ref {cref}")
                    cell_refs_seen.add(cref)
                    col_letters, row_in_ref = m.group(1), int(m.group(2))
                    if row_in_ref != rn:
                        issues.append(
                            f"{n}: cell {cref} placed inside <row r='{rn}'> "
                            f"(row component mismatches container)"
                        )
                    col_idx = _col_index(col_letters)
                    if cell_cols_seen and col_idx <= cell_cols_seen[-1]:
                        issues.append(
                            f"{n}: cell {cref} not in ascending column order in row {rn}"
                        )
                    cell_cols_seen.append(col_idx)

            # hyperlinks → location consistency
            hl_el = root.find(NS_M + "hyperlinks")
            if hl_el is not None:
                sheet_names_set = {s.get("name") for s in wb.findall(f".//{NS_M}sheet")} if wb is not None else set()
                for hl in hl_el.findall(NS_M + "hyperlink"):
                    ref = hl.get("ref", "")
                    loc = hl.get("location") or ""
                    if not cell_ref_re.match(ref):
                        issues.append(f"{n}: hyperlink ref '{ref}' is not a cell ref")
                    if loc and "!" in loc:
                        target_sheet = loc.split("!", 1)[0].strip("'")
                        if sheet_names_set and target_sheet not in sheet_names_set:
                            issues.append(
                                f"{n}: hyperlink {ref}→'{loc}' points at non-existent sheet "
                                f"'{target_sheet}'"
                            )

            # dimension sanity
            dim_el = root.find(NS_M + "dimension")
            if dim_el is not None and row_numbers_seen:
                dim_ref = dim_el.get("ref", "")
                if ":" in dim_ref:
                    last = dim_ref.split(":", 1)[1]
                    m = cell_ref_re.match(last)
                    if m and int(m.group(2)) < max(row_numbers_seen):
                        issues.append(
                            f"{n}: dimension ref='{dim_ref}' ends before last row "
                            f"r='{max(row_numbers_seen)}'"
                        )

        if wb is not None:
            sheets = wb.findall(f".//{NS_M}sheet")
            from collections import Counter

            for label, items in (
                ("sheet names", [s.get("name") for s in sheets]),
                ("sheetIds", [s.get("sheetId") for s in sheets]),
                ("rIds", [s.get(NS_REL_ + "id") for s in sheets]),
            ):
                dup = [x for x, c in Counter(items).items() if c > 1]
                if dup:
                    issues.append(f"Duplicate {label}: {dup}")

            rels_ids = {r.get("Id") for r in rels.findall(NS_PKG + "Relationship")}
            for s in sheets:
                rid = s.get(NS_REL_ + "id")
                if rid and rid not in rels_ids:
                    issues.append(f"sheet '{s.get('name')}' rId={rid} missing in rels")

            ct_parts = {o.get("PartName") for o in ct.findall(NS_CT_ + "Override")}
            for sheet_part in [n for n in names if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")]:
                if f"/{sheet_part}" not in ct_parts:
                    issues.append(f"sheet part /{sheet_part} has no ContentTypes Override")

    openpyxl_warnings: list[str] = []
    if through_openpyxl:
        import warnings as _w
        try:
            import openpyxl as _opx
            with _w.catch_warnings(record=True) as caught:
                _w.simplefilter("always")
                _opx.load_workbook(xlsx)
                for w in caught:
                    msg = f"{w.category.__name__}: {w.message}"
                    openpyxl_warnings.append(msg)
        except ImportError:
            typer.secho(
                "  ℹ openpyxl not installed — install with `pip install openpyxl` "
                "to enable --through-openpyxl",
                fg=typer.colors.YELLOW,
            )
        except Exception as e:
            openpyxl_warnings.append(f"load failed: {type(e).__name__}: {e}")

    typer.secho(f"\n=== Doctor report: {xlsx.name} ===", fg=typer.colors.CYAN, bold=True)
    typer.echo(f"  Sheets total      : {len(sheets) if wb is not None else '?'}")
    typer.echo(f"  XML parse errors  : {sum(1 for i in issues if 'parse error' in i)}")
    typer.echo(f"  Cells > 32767     : {len(over_limit)}")
    for n, L in over_limit[:10]:
        typer.secho(f"      {n}: {L} chars", fg=typer.colors.RED)
    typer.echo(f"  Bad control chars : {len(bad_chars)}")
    for n in bad_chars[:10]:
        typer.secho(f"      {n}", fg=typer.colors.RED)
    typer.echo(f"  Consistency issues: {len([i for i in issues if 'parse error' not in i])}")
    for i in issues[:30]:
        if "parse error" in i:
            continue
        typer.secho(f"      • {i[:300]}", fg=typer.colors.YELLOW)

    if through_openpyxl:
        typer.echo(f"  openpyxl warnings : {len(openpyxl_warnings)}")
        for w in openpyxl_warnings[:20]:
            typer.secho(f"      • {w[:300]}", fg=typer.colors.YELLOW)

    if not over_limit and not bad_chars and not issues and not openpyxl_warnings:
        typer.secho("\n✓ No defects detected.", fg=typer.colors.GREEN, bold=True)
    else:
        typer.secho(
            "\n⚠ Defects above explain Excel's «Ошибка в части содержимого» dialog.",
            fg=typer.colors.YELLOW, bold=True,
        )
        raise typer.Exit(1)


@app.command()
def build(
    project: Path = typer.Argument(..., help="Path to project.yaml."),
):
    """Build the Excel report according to project.yaml."""
    cfg, _ = _load_project(project)

    template_path = cfg.resolved_template()
    if not template_path.exists():
        typer.secho(f"Template not found: {template_path}", fg=typer.colors.RED, err=True)
        raise typer.Exit(1)
    if cfg.template is None:
        typer.secho(f"Using bundled clean template: {template_path}", fg=typer.colors.CYAN)

    catalog = Catalog.load_default()
    findings, unknown = _run_pipeline_for_domains(cfg, catalog)

    if not findings:
        typer.secho("No findings produced. Nothing to write.", fg=typer.colors.YELLOW)
        return

    cfg.output.parent.mkdir(parents=True, exist_ok=True)
    render_report(
        template_path,
        cfg.output,
        findings,
        clear_example_rows=cfg.defaults.clear_example_rows,
    )

    typer.secho(f"✓ Wrote {cfg.output}", fg=typer.colors.GREEN)
    typer.echo(f"  Findings: {len(findings)}")
    appendix_count = sum(1 for f in findings if f.appendix is not None)
    typer.echo(f"  Appendices: {appendix_count}")
    # Show per-domain breakdown if multi-domain
    domains_seen = {f.domain for f in findings if f.domain}
    if domains_seen:
        typer.echo(f"  Domains: {len(domains_seen)} — {', '.join(sorted(domains_seen))}")
    if unknown:
        typer.secho(
            f"⚠ {len(unknown)} unknown RiskId(s): {', '.join(unknown)} (written with placeholder)",
            fg=typer.colors.YELLOW,
        )


if __name__ == "__main__":
    app()
